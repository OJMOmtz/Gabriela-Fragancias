import os
import pyodbc
import pandas as pd
from openpyxl import load_workbook
import xlrd

# Lista de rutas de archivos Excel y Access
file_paths = [
    # Excel
    "D:\\Gabriela Fragancias\\2024\\Adidas Perfumes en 3NF_page001.xlsx",
    "D:\\Gabriela Fragancias\\2024\\pagares pendientes gabriela fragancias.xlsx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\CLIENTE.xls",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Clientes 2023 (2).xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Clientes 2023.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Clientes Cancelados (2).xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Clientes Cancelados.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Clientes Hist¢ricos.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Cobros zonas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Combined results(2023-08-04 213545).xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\INSERT.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\Listado de clientes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\2023\\habilitados y mesas por local padron2013_VERSION 6.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Dise¤os\\Clientes 2023.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\A Revisar (verificados).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\A Revisar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\MUESTRA PARA CARGA.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\Para SISCO.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\SISCO - A Ingresar 2020.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\SISCO - A Ingresar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\SISCO - A quitar (clientes ingresados por error).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\SISCO - Hoja 1 (clientes ya ingresados).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\SISCO - Saldo clientes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\InformConf\\Sacados.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Gabriela - Lista Perfumes 2022.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Joy Cat logo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Joy catalogo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Lares-Catalogo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Lista Multimarcas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Lista productos Silk (2022-01-19 00.52).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Lista productos belleza Silk (2022-01-19 00.52).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Ma Aux.accdb",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Morosos It .xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Morosos Oviedo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Morosos Santa Rita.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Morosos.accdb",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Morosos.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Ventas (Autoguardado).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Ventas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\Zonas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\base_fdb_PRODUCTO.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\base_fdb_PRODUCTO_MARCA.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\base_fdb_PRODUCTO_SEXO.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\base_fdb_PROVEEDOR.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\base_fdb_VENDEDOR.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\listado-.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\listado.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\multimarcas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Listas\\sairamperfumes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\090724_Silk_Lista_de_Precios_Base (2).xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\090724_Silk_Lista_de_Precios_Base.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\A Revisar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\BETO.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Control de Tarjetas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Cooperativa 26 de abril Ltda.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\DB Clientes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Detalle de cobro .xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\GABRIELA_CANCELADOS_-_FALTANTES.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Gabriela - Lista Perfumes 2022.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Joy Cat logo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Joy catalogo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lares-Catalogo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista Gabriel Fragancias - 2024.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista Gabriel Fragancias 2024.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista Multimarcas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista de precios.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista productos Silk (2022-01-19 00.52).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Lista productos belleza Silk (2022-01-19 00.52).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\MUESTRA PARA CARGA.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Ma Aux.accdb",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos It .xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos Oviedo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos Santa Rita.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos sin C‚dula.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos.accdb",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Morosos.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Para SISCO.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Proveedores y Compras.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\SISCO - A Ingresar 2020.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\SISCO - A Ingresar 2021.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\SISCO - A Ingresar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\SISCO - A quitar (clientes ingresados por error).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\SISCO - Hoja 1 (clientes ya ingresados).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Stock 13-08-2017.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Stock Maletines.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Stock Maletines.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Tablas normalizadas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Tarjetas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Ventas (Autoguardado).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Ventas.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Ventas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Zonas 2019 (2).xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Zonas 2022.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\Zonas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\base_fdb_PRODUCTO.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\base_fdb_PRODUCTO_MARCA.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\base_fdb_PRODUCTO_SEXO.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\base_fdb_PROVEEDOR.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\base_fdb_VENDEDOR.xls",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\listado-.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\listado.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\multimarcas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\Documentos\\Planillas\\sairamperfumes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\A Revisar (verificados).xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\A Revisar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Detalle de cobro .xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\GABRIELA_CANCELADOS_-_FALTANTES.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\MUESTRA PARA CARGA.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Morosos It .xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Morosos Oviedo.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Morosos Santa Rita.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Morosos.accdb",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Morosos.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Para SISCO.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - A Ingresar 2020.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - A Ingresar 2021.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - A Ingresar.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - A ingresar 2021-11-13.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - A quitar (clientes ingresados por error).xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - Hoja 1 (clientes ya ingresados).xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - Saldo clientes (2).xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\SISCO - Saldo clientes.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Sacados.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Ventas (Autoguardado).xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Ventas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\Zonas.xslx",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\base_fdb_PRODUCTO.xls",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\base_fdb_PRODUCTO_MARCA.xls",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\base_fdb_PRODUCTO_SEXO.xls",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\base_fdb_PROVEEDOR.xls",
    "D:\\Gabriela Fragancias\\Datos\\SISCO\\base_fdb_VENDEDOR.xls",
    "D:\\Gabriela Fragancias\\Programas\\Joiner\\data\\1_import.xslx",
    "D:\\Gabriela Fragancias\\Programas\\Joiner\\data\\2_import.xslx",
    "D:\\Gabriela Fragancias\\Programas\\Joiner\\data\\result.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\CLIENTE-20230817.xls",
    "D:\\Gabriela Fragancias\\preliminares\\Clientes 2023 (2).xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Clientes 2023 (3).xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Clientes 2023.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Clientes Hist¢ricos.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Gabriela Fragancias 2024 lista.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Lista De Perfumes.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Lista de precios.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Listado de clientes (2).xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Listado de clientes (3).xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Listado de clientes.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Morosos It .xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Morosos Oviedo.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Morosos Santa Rita.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Morosos.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Perfumes Y Olfato.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Ventas (Autoguardado).xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Ventas.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\Zonas.xslx",
    "D:\\Gabriela Fragancias\\preliminares\\base_fdb_PRODUCTO.xls",
    "D:\\Gabriela Fragancias\\preliminares\\base_fdb_PRODUCTO_MARCA.xls",
    "D:\\Gabriela Fragancias\\preliminares\\base_fdb_PRODUCTO_SEXO.xls",
    "D:\\Gabriela Fragancias\\preliminares\\base_fdb_PROVEEDOR.xls",
    "D:\\Gabriela Fragancias\\preliminares\\base_fdb_VENDEDOR.xls",
    
    # Access
    "D:\\Gabriela Fragancias\\Datos\\Ma Aux.accdb",
    "D:\\Gabriela Fragancias\\Datos\\Morosos.accdb",
    "F:\\ojmo\\ACER\\Vistumbler\\2023-08-09 15-46-03.mdb",
    "F:\\ojmo\\ACER\\Vistumbler\\2023-08-09 16-23-36.mdb",
    "F:\\ojmo\\Clientes\\Surgente\\Program Files\\Microsoft Office\\root\\Office16\\3082\\DBSAMPLE.mdb",
    "F:\\ojmo\\EZE\\Downloads\\Vistumbler\\2023-08-15 16-25-48.mdb",
    "F:\\ojmo\\EZE\\Downloads\\Vistumbler\\2023-08-15 16-52-32.mdb",
    "F:\\ojmo\\EZE\\Downloads\\Vistumbler\\2023-08-15 22-03-53.mdb",
    "F:\\ojmo\\Fonts\\#Colecciones\\56_hq_Fonts\\IDAutomation\\IDAutomation.com Free Code 39 Font\\Access 2000 Example.mdb",
    "F:\\ojmo\\Fonts\\#Colecciones\\56_hq_Fonts\\IDAutomation\\IDAutomation.com Free Code 39 Font\\Access 97 Example.mdb",
    "F:\\ojmo\\Fonts\\#Colecciones\\HQFonts\\IDAutomation\\IDAutomation.com Free Code 39 Font\\Access 2000 Example.mdb",
    "F:\\ojmo\\Fonts\\#Colecciones\\HQFonts\\IDAutomation\\IDAutomation.com Free Code 39 Font\\Access 97 Example.mdb",
    "F:\\ojmo\\PADRONES\\csvs\\CEDULA 2001.mdb",
    "F:\\ojmo\\PADRONES\\Epi_Info\\CDC.mdb",
    "F:\\ojmo\\PADRONES\\Pinturas\\Downloads\\renner industria\\Clientes Novos - Máquina Automátical\\Support\\RennerMM.mdb",
    "F:\\ojmo\\PADRONES\\Pinturas\\renner industria\\Clientes Novos - Máquina Automátical\\Support\\RennerMM.mdb",
    "F:\\ojmo\\PADRONES\\Pinturas\\RennerMM\\RennerMM.mdb",
    "F:\\ojmo\\SOFTWARE\\CODING\\BASES DE DATOS\\pentaho kettle-36e2662\\testfiles\\repository.mdb",
    "F:\\ojmo\\SOFTWARE\\CODING\\BASES DE DATOS\\pentaho-kettle-36e2662\\testfiles\\repository.mdb",
    "F:\\ojmo\\SOFTWARE\\CODING\\Logic Works\\ERwin 3.5\\Tutorial\\COMPARE.mdb",
    "F:\\ojmo\\SOFTWARE\\CODING\\Logic Works\\ERwin 3.5\\Tutorial\\reverse.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\Atlas2007_PY\\DataBases\\MyData.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\Atlas2007_PY\\DataBases\\MyData2005.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Albanian.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\ALBANIAN2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Arabic.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Arabic2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\BULGARIAN.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\BULGARIAN2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Czech.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Czech2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Danish.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Danish2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Dutch.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Dutch2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EF-FEdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EG-GEdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EI-IEdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EN-CSdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EN-CTdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\EP-PEdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\ES-SEdata.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Esperanto.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Esperanto2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Farsi.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Finnish.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Finnish2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Greek.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Greek2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Hebrew.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Hebrew2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\HUNGARIAN.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\HUNGARIAN2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Indonesian.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Indonesian2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Korean.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Language.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Latin.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Latin2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Latvian.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Latvian2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\NORVEGIAN.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Polish.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Polish2.mdb",
    "F:\\ojmo\\SOFTWARE\\EDUCATIVOS\\L&H Power Translator Pro\\WLData\\Romanian.mdb",
    "F:\\ojmo\\SOFTWARE\\PRODUCTIVIDAD\\SAm\\SolidWorks Data\\lang\\english\\SWBrowser.mdb",
]

# Procesar cada archivo en la lista de rutas
for file_path in file_paths:
    if os.path.isfile(file_path):
        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == ".xlsx":
            # Leer archivo Excel (.xlsx) usando openpyxl
            workbook = load_workbook(file_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data = sheet.values
                try:
                    columns = next(data)[0:]
                    df = pd.DataFrame(data, columns=columns)
                    print(f"\nDatos del archivo Excel '{file_path}', hoja '{sheet_name}':")
                    print(df)
                except StopIteration:
                    print(f"\nLa hoja '{sheet_name}' en el archivo Excel '{file_path}' está vacía.")

        elif file_extension == ".xls":
            # Leer archivo Excel (.xls) usando xlrd
            workbook = xlrd.open_workbook(file_path)
            for sheet_name in workbook.sheet_names():
                sheet = workbook.sheet_by_name(sheet_name)
                data = [sheet.row_values(rowx) for rowx in range(sheet.nrows)]
                if len(data) > 1:
                    columns = data[0][0:]
                    df = pd.DataFrame(data[1:], columns=columns)
                    print(f"\nDatos del archivo Excel '{file_path}', hoja '{sheet_name}':")
                    print(df)
                else:
                    print(f"\nLa hoja '{sheet_name}' en el archivo Excel '{file_path}' está vacía.")

        elif file_extension == ".mdb" or file_extension == ".accdb":
            # Leer archivo Access
            conn_str = (
                r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
                f'DBQ={file_path};'
            )
            conn = pyodbc.connect(conn_str)
            cursor = conn.cursor()

            for table_info in cursor.tables(tableType='TABLE'):
                table_name = table_info.table_name
                df = pd.read_sql(f"SELECT * FROM [{table_name}]", conn)
                print(f"\nDatos de la tabla '{table_name}' en el archivo Access '{file_path}':")
                print(df)

            conn.close()

        else:
            print(f"\nEl archivo '{file_path}' no es un formato compatible.")

    else:
        print(f"\nEl archivo '{file_path}' no existe.")
